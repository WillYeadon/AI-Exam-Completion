import re
from openpyxl import load_workbook

def rename_values_in_cells(excel_file_path):
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=False):
        cell = row[0]
        file_path = cell.value
        match = re.search(r'(\d{4})-papers-raw\\(.+?)~Q\d+\.tex$', file_path)
        if match:
            year, title = match.groups()
            new_value = f'{year}-{title}'
            cell.value = new_value
            print(f'Renamed: {file_path} -> {new_value}')

    workbook.save('2019-2017-checked.xlsx')

if __name__ == '__main__':
    excel_file_path = '2019-2017-checked.xlsx'
    rename_values_in_cells(excel_file_path)
